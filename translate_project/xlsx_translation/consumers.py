from channels.generic.websocket import WebsocketConsumer
import json, os, shutil
from time import sleep
from random import randint
from googletrans import Translator
import pandas as pd
from pathlib import Path
import os, shutil
import editpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import glob



def get_latest_file():
    directory="./media/excel/"
    
    list_of_files = glob.glob(f'{directory}/*')  # Get all files in the directory
    if not list_of_files:
        return None
    latest_file = max(list_of_files, key=os.path.getctime)  # Get the latest file based on creation time
    return latest_file




def is_english(text):
    return all(ord(c) < 128 for c in text)



class progressConsumer(WebsocketConsumer):
    
        
    def connect(self):
        self.accept()  
        
         
            
            
    def disconnect(self, close_code):
        pass
    
    
    
    def receive(self, text_data):
                
        
        
        data = json.loads(text_data)
        action = data.get('action')
        lng=data.get('language')
        
        print("Language selected", lng)

        if action == 'button_clicked':
            
                
            
                print("Translation is in progress....")
                
                try:
                    
                
                    src_file_path=get_latest_file()
                    file_name=src_file_path.split("\\")[1].split(".")[0]
                    file_type=src_file_path.split("\\")[1].split(".")[1]
                    
                    if file_type!='xlsx':
                        
                        self.send(json.dumps({'value': 3.1425}))
                        return None
                    
                
                except:
                    
                    self.send(json.dumps({'value': 3.142}))
                    return None
                
                               
                
            
                translator = Translator()
                

                #copy the file to the destination folder, the make a copy of the desired language
                if os.path.exists(f"./converted_data/{file_name}.xlsx"):
                    pass
                else:
                    shutil.copy(src_file_path, f"./converted_data/{file_name}.xlsx")


                if os.path.exists(f"./converted_data/{file_name}_{lng}.xlsx"):
                    pass
                else:
                    shutil.copy(src_file_path, f"./converted_data/{file_name}_{lng}.xlsx")

                #Start the language transaltion
                src_wkbk = f'./converted_data/{file_name}.xlsx'
            
                dest_wkbk = f"./converted_data/{file_name}_{lng}.xlsx"
                
                ###Later make it dynamic
                
                
                if lng=='hindi':
                    conv_lng='hi'
                elif lng=='gujarati':
                    conv_lng='gu'
                else:
                    conv_lng='en'

                wb = editpyxl.Workbook() 
                
                
                
                wb.open(src_wkbk)
                sheets = wb.sheetnames
                
                self.send(json.dumps({'value': 1}))
                
                    #Iterate sheet over sheet
                for count, sheet in enumerate(sheets):
                    
                            
                    #Udpate the progress bar
                    print(count)
                    
                   
                    
                    #Get the Start letter, end letter, start row, end row, ofr each sheet
                    wb_range = load_workbook(src_wkbk)
                    ws_range=wb_range[sheet]
                    
                    min_row=ws_range.min_row
                    max_row=ws_range.max_row
                    min_col=ws_range.min_column
                    max_col=ws_range.max_column
                    
                    col_list = [get_column_letter(col) for col in range(1, max_col + 1)]
                    row_range=list(range(min_row, max_row+1))
                    
                    
                    
                    wb_range.close()
                        
                    ws = wb[sheet]
                    
                    
                    #Iterate Cell by cell and cpopy the respective contents
                    for col in col_list:

                        for rw in row_range:

                            try:

                                if ws.cell(f"{col}{rw}").value!=None:
                                    
                                    
                                    if is_english(ws.cell(f"{col}{rw}").value):
                                        

                                        ws[f"{col}{rw}"] =  translator.translate(ws.cell(f"{col}{rw}").value , src='en', dest=conv_lng).text
                                    
                                    


                            except:
                                pass
                      
                    #Update the progress bar        
                    self.send(json.dumps({'value': round( ( (count+1)/ len(sheets) )*100, 2  )}))        

                    
                    
                self.send(json.dumps({'value': 100 }))
                wb.save(dest_wkbk)
                
                wb.close()
                print("Translation Complete")
                
                #With some delay reset the progress bar
                sleep(6)
                self.send(json.dumps({'value': 0 }))
                self.send(json.dumps({'value': 3.14256 }))
                
                #empty the source directory once the conversion is complete
                directory="./media/excel/"
                list_of_files = glob.glob(f'{directory}/*')
                
                for file in list_of_files:
                    os.remove(file)
                
                
                return 'Done'


            
            
            
                       
            
 
            
            
            
            
            
            
          
            
        