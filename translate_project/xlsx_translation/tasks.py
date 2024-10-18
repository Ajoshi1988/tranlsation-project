from celery import shared_task
from time import sleep
from celery_progress.backend import ProgressRecorder
from googletrans import Translator
import pandas as pd
from pathlib import Path
import os, shutil
import editpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



###########################################################################################



def is_english(text):
    return all(ord(c) < 128 for c in text)


def convert_the_data(self, file_name, lng):
    
    
    
    progress_recorder=ProgressRecorder(self)

    translator = Translator()
    #create an folder if the one doest not exists
    directory = Path(f"./converted_data/{file_name}")

    #check for the path, create the folder with same name as filename,
    if not directory.exists():
        directory.mkdir(parents=True, exist_ok=True)
        print(f"Directory '{directory}' created successfully!")
    else:
        print(f"Directory '{directory}' already exists.")

    #copy the file to the destination folder, the make a copy of the desired language
    if os.path.exists(f"./converted_data/{file_name}/{file_name}.xlsx"):
        pass
    else:
        shutil.copy(f"{file_name}.xlsx", f"./converted_data/{file_name}/{file_name}.xlsx")


    if os.path.exists(f"./converted_data/{file_name}/{file_name}_{lng}.xlsx"):
        pass
    else:
        shutil.copy(f"{file_name}.xlsx", f"./converted_data/{file_name}/{file_name}_{lng}.xlsx")

    #Start the language transaltion
    src_wkbk = f'./converted_data/{file_name}/{file_name}.xlsx'
   
    dest_wkbk = f"./converted_data/{file_name}/{file_name}_{lng}.xlsx"
       

    
    if lng=='hindi':
        conv_lng='hi'
    elif lng=='gujarati':
        conv_lng='gu'
    else:
        conv_lng='en'

    wb = editpyxl.Workbook() 
    
    
    
    wb.open(src_wkbk)
    sheets = wb.sheetnames
   
    
    
    
    #Iterate sheet over sheet
    for count, sheet in enumerate(sheets):
        
                  
        #Udpate the progress bar
        progress_recorder.set_progress(count,len(sheets),f'Processing Sheet {sheet}')
        
        #Get the Start letter, end letter, start row, end row, ofr each sheet
        wb_range = load_workbook(src_wkbk)
        ws_range=wb_range[sheet]
        
        min_row=ws_range.min_row
        max_row=ws_range.max_row
        min_col=ws_range.min_column
        max_col=ws_range.max_column
        
        col_list = [get_column_letter(col) for col in range(1, max_col + 1)]
        row_range=list(range(min_row, max_row+1))
        
        
        
        wb_range.close()
              
        ws = wb[sheet]
        
        
        #Iterate Cell by cell and cpopy the respective contents
        for col in col_list:

            for rw in row_range:

                try:

                    if ws.cell(f"{col}{rw}").value!=None:
                        
                        
                        if is_english(ws.cell(f"{col}{rw}").value):
                            

                            ws[f"{col}{rw}"] =  translator.translate(ws.cell(f"{col}{rw}").value , src='en', dest=conv_lng).text
                        
                        


                except:
                    pass

       
        
    
    wb.save(dest_wkbk)
      
    wb.close()
    
    return 'Done'
    




